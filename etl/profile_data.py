"""
DE-4: Profile and assess data quality of all Excel sources.

For each sheet, reports:
  - Null counts per column
  - Duplicate rows
  - Detected column types
  - Schema inconsistencies (e.g., mixed types in a column)

Unchanged files are skipped automatically — only new or changed files
are re-profiled. Output is written to data/profiles/data_quality_report.json
and a human-readable summary is printed to stdout.

Usage:
    python -m etl.profile_data             # skip unchanged files
    python -m etl.profile_data --force     # re-profile all files
"""

import json
import sys
from collections import Counter
from datetime import datetime, timezone

import openpyxl

from config import PROCESSED_DIR, PROFILES_DIR, RAW_DATA_DIR, REPORT_PATH, ROOT, SOURCES, TRACKER_PATH
from etl.file_tracker import FileTracker


def detect_type(value) -> str:
    if value is None:
        return "null"
    if isinstance(value, bool):
        return "bool"
    if isinstance(value, int):
        return "int"
    if isinstance(value, float):
        return "float"
    if isinstance(value, datetime):
        return "datetime"
    if isinstance(value, str):
        if value.strip().startswith("="):
            return "formula"
        return "str"
    return type(value).__name__


def find_header_row(rows: list) -> int:
    """Return 0-based index of the first row with >= 2 non-None cells."""
    for i, row in enumerate(rows):
        if len([c for c in row if c is not None]) >= 2:
            return i
    return 0


def profile_sheet(ws) -> dict:
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return {"error": "empty sheet"}

    max_cols = max((len(r) for r in rows), default=0)
    rows_padded = [list(r) + [None] * (max_cols - len(r)) for r in rows]

    header_idx = find_header_row(rows_padded)
    header = [str(c).strip() if c is not None else f"col_{i}" for i, c in enumerate(rows_padded[header_idx])]
    data_rows = rows_padded[header_idx + 1:]

    total_data = len(data_rows)
    non_empty_data = sum(1 for r in data_rows if any(c is not None for c in r))

    columns = []
    for col_idx, col_name in enumerate(header):
        values = [r[col_idx] if col_idx < len(r) else None for r in data_rows]
        non_null_values = [v for v in values if v is not None]
        null_count = len(values) - len(non_null_values)
        type_counts = Counter(detect_type(v) for v in non_null_values)
        formula_count = type_counts.pop("formula", 0)
        inconsistent = len(type_counts) > 1

        columns.append({
            "column": col_name,
            "null_count": null_count,
            "null_pct": round(null_count / len(values) * 100, 1) if values else 0,
            "non_null_count": len(non_null_values),
            "dominant_type": type_counts.most_common(1)[0][0] if type_counts else "unknown",
            "type_distribution": dict(type_counts),
            "formula_cells": formula_count,
            "type_inconsistent": inconsistent,
        })

    row_strings = [str(r) for r in data_rows if any(c is not None for c in r)]
    dup_counts = Counter(row_strings)
    duplicate_row_count = sum(v - 1 for v in dup_counts.values() if v > 1)

    issues = []
    for col in columns:
        if col["type_inconsistent"]:
            issues.append(f"Column '{col['column']}': mixed types {col['type_distribution']}")
        if col["null_pct"] > 50:
            issues.append(f"Column '{col['column']}': {col['null_pct']}% null")
        if col["formula_cells"] > 0:
            issues.append(f"Column '{col['column']}': {col['formula_cells']} formula cell(s) (data_only=True may suppress values)")
    if duplicate_row_count > 0:
        issues.append(f"{duplicate_row_count} duplicate data row(s) detected")

    return {
        "header_row": header_idx + 1,
        "column_headers": header,
        "total_data_rows": total_data,
        "non_empty_data_rows": non_empty_data,
        "duplicate_row_count": duplicate_row_count,
        "columns": columns,
        "issues": issues,
        "issue_count": len(issues),
    }


def profile_file(filename: str, source_id: str) -> dict:
    path = RAW_DATA_DIR / filename
    if not path.exists():
        return {"source_id": source_id, "filename": filename, "error": "file not found", "sheets": []}

    try:
        wb = openpyxl.load_workbook(str(path), data_only=True)
    except Exception as exc:
        return {"source_id": source_id, "filename": filename, "error": str(exc), "sheets": []}

    sheet_profiles = []
    total_issues = 0
    for name in wb.sheetnames:
        ws = wb[name]
        prof = profile_sheet(ws)
        prof["sheet_name"] = name
        sheet_profiles.append(prof)
        total_issues += prof.get("issue_count", 0)

    return {
        "source_id": source_id,
        "filename": filename,
        "sheet_count": len(sheet_profiles),
        "total_issues": total_issues,
        "sheets": sheet_profiles,
    }


def load_existing_report() -> dict:
    """Load existing report entries keyed by source_id (for merging skipped files)."""
    if not REPORT_PATH.exists():
        return {}
    with open(REPORT_PATH, encoding="utf-8") as f:
        data = json.load(f)
    return {s["source_id"]: s for s in data.get("sources", [])}


def print_summary(report: dict):
    print("\n" + "=" * 60)
    print("DATA QUALITY REPORT SUMMARY")
    print("=" * 60)
    for src in report["sources"]:
        print(f"\n[{src['source_id']}] {src['filename']}")
        if "error" in src:
            print(f"  ERROR: {src['error']}")
            continue
        print(f"  Sheets: {src['sheet_count']}  |  Total issues: {src['total_issues']}")
        for sheet in src["sheets"]:
            issue_count = sheet.get("issue_count", 0)
            flag = "OK" if issue_count == 0 else f"! {issue_count} issue(s)"
            print(f"  [{flag:18s}] {sheet['sheet_name']}  ({sheet.get('non_empty_data_rows', '?')} data rows)")
            for issue in sheet.get("issues", []):
                print(f"                     >> {issue}")


def main():
    force = "--force" in sys.argv
    tracker = FileTracker(TRACKER_PATH, ROOT)
    existing = load_existing_report()

    PROFILES_DIR.mkdir(parents=True, exist_ok=True)
    PROCESSED_DIR.mkdir(parents=True, exist_ok=True)

    report = {
        "generated_at": datetime.now(timezone.utc).isoformat(),
        "sources": [],
    }

    skipped = profiled = 0
    for source in SOURCES:
        filename, source_id = source["filename"], source["source_id"]
        filepath = RAW_DATA_DIR / filename

        if not force and not tracker.needs_processing(filepath):
            entry = existing.get(source_id)
            if entry:
                report["sources"].append(entry)
                status = tracker.get_status(filepath)
                print(f"Skipping  {filename} — unchanged (status: {status})")
                skipped += 1
                continue

        print(f"Profiling {filename} ...", end=" ", flush=True)
        result = profile_file(filename, source_id)
        report["sources"].append(result)
        issues = result.get("total_issues", "?")
        print(f"done ({issues} issue(s))")

        if "error" not in result:
            # Write individual processed profile
            out = PROCESSED_DIR / f"{source_id}_profile.json"
            with open(out, "w", encoding="utf-8") as f:
                json.dump(result, f, indent=2, default=str)
            # Only advance to processed if not already loaded
            if tracker.get_status(filepath) != "loaded":
                tracker.mark_processed(filepath, source_id=source_id)
        profiled += 1

    with open(REPORT_PATH, "w", encoding="utf-8") as f:
        json.dump(report, f, indent=2, default=str)

    print_summary(report)
    print(f"\nFull report written to {REPORT_PATH}")
    print(f"  Profiled: {profiled}  |  Skipped (unchanged): {skipped}")


if __name__ == "__main__":
    main()
