"""
DE-3: Ingest and catalog all client Excel datasets.

Loads each Excel file from Raw_data/, extracts sheet-level metadata,
and writes a structured catalog to data/catalog/catalog.json.

Unchanged files are skipped automatically — only new or changed files
are re-cataloged. Each processed file also gets an individual entry
written to data/processed/{source_id}_catalog.json.
Raw files are never modified.

Usage:
    python -m etl.ingest_catalog             # skip unchanged files
    python -m etl.ingest_catalog --force     # re-catalog all files
"""

import json
import sys
from datetime import datetime, timezone

import openpyxl

from config import CATALOG_DIR, CATALOG_PATH, PROCESSED_DIR, RAW_DATA_DIR, ROOT, SOURCES, TRACKER_PATH
from etl.file_tracker import FileTracker


def profile_sheet(ws) -> dict:
    """Return basic metadata for a single worksheet."""
    rows = list(ws.iter_rows(values_only=True))
    total_rows = len(rows)
    non_empty_rows = sum(1 for r in rows if any(c is not None for c in r))
    col_count = ws.max_column or 0

    header_row = None
    for i, row in enumerate(rows):
        vals = [str(v).strip() for v in row if v is not None]
        if len(vals) >= 2:
            header_row = i + 1
            break

    return {
        "total_rows": total_rows,
        "non_empty_rows": non_empty_rows,
        "column_count": col_count,
        "header_row": header_row,
    }


def catalog_file(source: dict) -> dict:
    path = RAW_DATA_DIR / source["filename"]
    if not path.exists():
        return {
            **source,
            "status": "missing",
            "error": f"File not found: {path}",
            "sheets": [],
        }

    try:
        wb = openpyxl.load_workbook(str(path), data_only=True, read_only=True)
    except Exception as exc:
        return {
            **source,
            "status": "error",
            "error": str(exc),
            "sheets": [],
        }

    sheets = []
    for name in wb.sheetnames:
        ws = wb[name]
        meta = profile_sheet(ws)
        sheets.append({"sheet_name": name, **meta})

    wb.close()

    return {
        **source,
        "status": "ok",
        "path": str(path.relative_to(ROOT)),
        "file_size_kb": round(path.stat().st_size / 1024, 1),
        "sheet_count": len(sheets),
        "sheets": sheets,
    }


def load_existing_catalog() -> dict:
    """Load existing catalog entries keyed by source_id (for merging skipped files)."""
    if not CATALOG_PATH.exists():
        return {}
    with open(CATALOG_PATH, encoding="utf-8") as f:
        data = json.load(f)
    return {s["source_id"]: s for s in data.get("sources", [])}


def main():
    force = "--force" in sys.argv
    tracker = FileTracker(TRACKER_PATH, ROOT)
    existing = load_existing_catalog()

    CATALOG_DIR.mkdir(parents=True, exist_ok=True)
    PROCESSED_DIR.mkdir(parents=True, exist_ok=True)

    catalog = {
        "generated_at": datetime.now(timezone.utc).isoformat(),
        "raw_data_dir": str(RAW_DATA_DIR.relative_to(ROOT)),
        "sources": [],
    }

    skipped = processed = 0
    for source in SOURCES:
        filepath = RAW_DATA_DIR / source["filename"]

        if not force and not tracker.needs_processing(filepath):
            entry = existing.get(source["source_id"])
            if entry:
                catalog["sources"].append(entry)
                status = tracker.get_status(filepath)
                print(f"Skipping  {source['filename']} — unchanged (status: {status})")
                skipped += 1
                continue

        print(f"Cataloging {source['filename']} ...", end=" ", flush=True)
        entry = catalog_file(source)
        catalog["sources"].append(entry)

        if entry["status"] == "ok":
            print(f"OK ({entry['sheet_count']} sheets, {entry['file_size_kb']} KB)")
            # Write individual processed output
            out = PROCESSED_DIR / f"{source['source_id']}_catalog.json"
            with open(out, "w", encoding="utf-8") as f:
                json.dump(entry, f, indent=2, default=str)
            tracker.mark_processed(filepath, source_id=source["source_id"])
            processed += 1
        else:
            print(f"FAIL — {entry.get('error')}")

    with open(CATALOG_PATH, "w", encoding="utf-8") as f:
        json.dump(catalog, f, indent=2, default=str)

    print(f"\nCatalog written to {CATALOG_PATH}")
    print(f"  Processed: {processed}  |  Skipped (unchanged): {skipped}")


if __name__ == "__main__":
    main()
