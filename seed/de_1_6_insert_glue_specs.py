"""
DE-1.6: Parse spec strings and insert properties for all glue_ledger sources.

For each source in config.SOURCES with seed_type="glue_ledger":
  1. Read the Specifications column (first data row).
  2. Parse the format "<weight>kg/<count>bag" (e.g. "25kg/1bag").
  3. Insert item_property rows for the matching material:
       weight_per_bag_kg  — numeric weight in kg
       bags_per_unit      — number of bags per purchasing unit

To add a new adhesive source, append one entry to SOURCES in config.py
with seed_type="glue_ledger" — no changes needed here.

Failure handling:
  - Material not in DB → blocked on DE-1.5.
  - Spec column empty → logged as failure.
  - Unrecognised format → logged; does not stop other sources.
  - Property already present → skipped (idempotent).

Idempotency key: "<source_id>.<property_name>"
  e.g. "glue_record.weight_per_bag_kg"

Usage:
    python -m seed.de_1_6_insert_glue_specs
    python -m seed.de_1_6_insert_glue_specs --retry
"""

import argparse
import json
import re
import sys
from datetime import datetime, timezone
from decimal import Decimal, InvalidOperation

import openpyxl

from config import LOADED_DIR, RAW_DATA_DIR, ROOT, SOURCES, TRACKER_PATH
from etl.file_tracker import FileTracker
from seed.db import get_connection
from seed.seed_logger import SeedLogger

SCRIPT_NAME  = "de_1_6"
COL_ITEMS    = 1
COL_SPECS    = 2
PROPERTIES   = ("weight_per_bag_kg", "bags_per_unit")

SPEC_RE = re.compile(r"^(\d+(?:\.\d+)?)kg/(\d+)bag$", re.IGNORECASE)

GLUE_SOURCES = [s for s in SOURCES if s.get("seed_type") == "glue_ledger"]


# ------------------------------------------------------------------ #
# Workbook helpers                                                     #
# ------------------------------------------------------------------ #

def find_header_row_idx(ws) -> int:
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        vals = [str(v).strip().lower() for v in row if v is not None]
        if "items" in vals and "specifications" in vals:
            return i
    return -1


def get_display_name(ws, header_idx: int):
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i <= header_idx:
            continue
        val = row[COL_ITEMS] if len(row) > COL_ITEMS else None
        if val is not None and str(val).strip():
            return str(val).strip().title()
    return None


def collect_spec_strings(ws, header_idx: int) -> list:
    results = []
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i <= header_idx:
            continue
        val = row[COL_SPECS] if len(row) > COL_SPECS else None
        if val is not None and str(val).strip():
            results.append((i + 1, str(val).strip()))
    return results


def parse_spec(raw: str):
    m = SPEC_RE.match(raw.strip())
    if not m:
        return None, f"Unrecognised format '{raw}' — expected e.g. '25kg/1bag'"
    try:
        weight = Decimal(m.group(1))
        bags   = Decimal(m.group(2))
    except InvalidOperation as exc:
        return None, f"Non-numeric value in '{raw}': {exc}"
    return (weight, bags), None


# ------------------------------------------------------------------ #
# Database helpers                                                     #
# ------------------------------------------------------------------ #

def db_get_material(conn, display_name: str):
    with conn.cursor() as cur:
        cur.execute("SELECT id FROM inventory_item WHERE display_name = %s", (display_name,))
        row = cur.fetchone()
    return row[0] if row else None


def db_property_exists(conn, item_id: int, prop_name: str) -> bool:
    with conn.cursor() as cur:
        cur.execute(
            "SELECT 1 FROM item_property WHERE item_id = %s AND property_name = %s",
            (item_id, prop_name),
        )
        return cur.fetchone() is not None


def db_insert_property(conn, item_id: int, prop_name: str, value: Decimal):
    with conn.cursor() as cur:
        cur.execute(
            "INSERT INTO item_property (item_id, property_name, property_value) VALUES (%s, %s, %s)",
            (item_id, prop_name, float(value)),
        )
    conn.commit()


# ------------------------------------------------------------------ #
# Per-source processing                                                #
# ------------------------------------------------------------------ #

def process_source(source: dict, conn, logger: SeedLogger, is_retry: bool, failed_keys: set) -> dict:
    """Process one glue_ledger source. Returns stats dict."""
    source_id = source["source_id"]
    filepath  = RAW_DATA_DIR / source["filename"]
    stats = {"source_id": source_id, "inserted": 0, "skipped": 0, "failed": 0}

    try:
        wb = openpyxl.load_workbook(str(filepath), data_only=True, read_only=True)
    except Exception as exc:
        key = f"{source_id}.all"
        logger.log_failure(key, f"Cannot open workbook: {exc}", str(filepath))
        print(f"  FAIL  [{source['filename']}] Cannot open: {exc}")
        stats["failed"] += len(PROPERTIES)
        return stats

    ws = wb.worksheets[0]
    header_idx = find_header_row_idx(ws)
    if header_idx < 0:
        wb.close()
        key = f"{source_id}.all"
        logger.log_failure(key, "Header row not found", f"sheet={ws.title}")
        print(f"  FAIL  [{source['filename']}] Header row not found")
        stats["failed"] += len(PROPERTIES)
        return stats

    display_name = get_display_name(ws, header_idx)
    spec_rows    = collect_spec_strings(ws, header_idx)
    wb.close()

    if display_name is None:
        key = f"{source_id}.all"
        logger.log_failure(key, "Items column empty", f"file={source['filename']}")
        print(f"  FAIL  [{source['filename']}] Items column empty")
        stats["failed"] += len(PROPERTIES)
        return stats

    item_id = db_get_material(conn, display_name)
    if item_id is None:
        key = f"{source_id}.all"
        logger.log_failure(key, f"'{display_name}' not in DB — blocked on DE-1.5", "")
        print(f"  BLOCK [{source['filename']}] '{display_name}' not in DB — run DE-1.5 first")
        stats["failed"] += len(PROPERTIES)
        return stats

    if not spec_rows:
        key = f"{source_id}.all"
        logger.log_failure(key, "Specifications column empty", f"file={source['filename']}")
        print(f"  FAIL  [{source['filename']}] Specifications column empty")
        stats["failed"] += len(PROPERTIES)
        return stats

    distinct = list(dict.fromkeys(s for _, s in spec_rows))
    if len(distinct) > 1:
        print(f"  WARN  [{source['filename']}] {len(distinct)} distinct specs: {distinct}")

    canonical_row, canonical_raw = spec_rows[0]
    parsed, err = parse_spec(canonical_raw)
    if err:
        key = f"{source_id}.parse"
        logger.log_failure(key, err, f"row={canonical_row}, raw='{canonical_raw}'")
        print(f"  FAIL  [{source['filename']}] {err}")
        stats["failed"] += len(PROPERTIES)
        return stats

    weight_per_bag_kg, bags_per_unit = parsed
    print(f"  [{source['filename']}] '{display_name}' — spec='{canonical_raw}'")

    for prop_name, prop_value in zip(PROPERTIES, (weight_per_bag_kg, bags_per_unit)):
        retry_key = f"{source_id}.{prop_name}"

        if is_retry and retry_key not in failed_keys:
            continue

        if db_property_exists(conn, item_id, prop_name):
            logger.log_success(retry_key)
            print(f"    SKIP  {prop_name}={prop_value} already in DB")
            stats["skipped"] += 1
            continue

        try:
            db_insert_property(conn, item_id, prop_name, prop_value)
            logger.log_success(retry_key)
            print(f"    INSERT {prop_name}={prop_value}")
            stats["inserted"] += 1
        except Exception as exc:
            conn.rollback()
            logger.log_failure(retry_key, f"DB insert failed: {exc}", f"item_id={item_id}")
            print(f"    FAIL  {prop_name}: {exc}")
            stats["failed"] += 1

    return stats


# ------------------------------------------------------------------ #
# Entry point                                                          #
# ------------------------------------------------------------------ #

def main():
    parser = argparse.ArgumentParser(description="DE-1.6: Insert spec properties for all glue_ledger sources")
    parser.add_argument("--retry", action="store_true")
    args = parser.parse_args()

    if not GLUE_SOURCES:
        print("No glue_ledger sources configured in config.SOURCES — nothing to do.")
        sys.exit(0)

    logger = SeedLogger(SCRIPT_NAME)
    failed_keys = logger.get_failed_keys()
    is_retry = args.retry or bool(failed_keys)

    print(f"{'RETRY' if is_retry else 'FULL'} mode — {len(GLUE_SOURCES)} glue_ledger source(s)\n")

    conn = get_connection()
    tracker = FileTracker(TRACKER_PATH, ROOT)
    all_stats = []
    total_failed = 0

    for source in GLUE_SOURCES:
        stats = process_source(source, conn, logger, is_retry, failed_keys)
        all_stats.append(stats)
        total_failed += stats["failed"]
        if stats["failed"] == 0:
            tracker.mark_loaded(RAW_DATA_DIR / source["filename"])

    conn.close()

    print("\n" + "=" * 52)
    print("DE-1.6 SUMMARY")
    for s in all_stats:
        print(f"  {s['source_id']:<22} inserted={s['inserted']}  skipped={s['skipped']}  failed={s['failed']}")
    print(f"\n  Total failed: {total_failed}")
    print("=" * 52)

    if not total_failed:
        LOADED_DIR.mkdir(parents=True, exist_ok=True)
        manifest = {
            "script": SCRIPT_NAME,
            "sources_processed": [s["source_id"] for s in GLUE_SOURCES],
            "mode": "retry" if is_retry else "full",
            "loaded_at": datetime.now(timezone.utc).isoformat(),
            "summary": all_stats,
        }
        with open(LOADED_DIR / "glue_specs_de_1_6.json", "w", encoding="utf-8") as f:
            json.dump(manifest, f, indent=2)
        print(f"  Load manifest: {LOADED_DIR / 'glue_specs_de_1_6.json'}")

    sys.exit(1 if total_failed else 0)


if __name__ == "__main__":
    main()
