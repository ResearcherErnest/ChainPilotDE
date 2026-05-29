"""
DE-1.1: Parse Stock_management.xlsx tabs and insert board material records.

For each tab whose name matches a thickness pattern (e.g. '6MM', '9MM'),
one inventory_item is inserted with:
    display_name = <Items-column value> + ' MDF'
    category     = 'Raw Material'
    base_uom     = 'sheets (pcs)'

Idempotency rules:
  - Fully complete record already in DB → skip.
  - Partially written record in DB → repair (fill missing fields).
  - Name collision from a different source tab → logged as failure.

Retry mode (auto-detected or --retry flag):
  - When seed_failures.log contains entries for de_1_1, only those tabs
    are processed. All other valid tabs are skipped as already-done.

Usage:
    python -m seed.de_1_1_insert_materials
    python -m seed.de_1_1_insert_materials --retry
"""

import argparse
import os
import re
import sys

import openpyxl

sys.path.insert(0, os.path.dirname(os.path.dirname(__file__)))
from seed.db import get_connection
from seed.seed_logger import SeedLogger

SCRIPT_NAME = "de_1_1"
WORKBOOK_PATH = os.path.join(
    os.path.dirname(os.path.dirname(__file__)), "Raw_data", "Stock management.xlsx"
)

# Matches tab names like '6MM', '9MM', '35MM ', '14mm' — ignores 'sales form'
THICKNESS_RE = re.compile(r"^\d+\s*[Mm]{2}\s*$")

# Column indices (0-based) within each sheet
COL_ITEMS = 1   # 'Items' (column B)


# ------------------------------------------------------------------ #
# Workbook helpers                                                     #
# ------------------------------------------------------------------ #

def is_valid_tab(name: str) -> bool:
    return bool(THICKNESS_RE.match(name))


def find_header_row_idx(ws) -> int:
    """Return 0-based index of the row containing 'Items' and 'Date'."""
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        vals = [str(v).strip().lower() for v in row if v is not None]
        if "items" in vals and "date" in vals:
            return i
    return -1


def get_first_data_row(ws, header_idx: int):
    """
    Return (row_tuple, 1-based-row-number) for the first row after the header
    that has a non-empty Items value. Returns (None, None) if not found.
    """
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i <= header_idx:
            continue
        items_val = row[COL_ITEMS] if len(row) > COL_ITEMS else None
        if items_val is not None and str(items_val).strip():
            return row, i + 1
    return None, None


# ------------------------------------------------------------------ #
# Database helpers                                                     #
# ------------------------------------------------------------------ #

CATEGORY = "Raw Material"
BASE_UOM = "sheets (pcs)"


def db_check_material(conn, display_name: str):
    """
    Returns (item_id, is_complete) where is_complete means
    both category and base_uom are non-null.
    Returns (None, False) when the record does not exist.
    """
    with conn.cursor() as cur:
        cur.execute(
            "SELECT id, category, base_uom FROM inventory_item WHERE display_name = %s",
            (display_name,),
        )
        row = cur.fetchone()
    if row is None:
        return None, False
    item_id, category, base_uom = row
    return item_id, (category is not None and base_uom is not None)


def db_insert_material(conn, display_name: str) -> int:
    """Insert a new inventory_item and return its id."""
    with conn.cursor() as cur:
        cur.execute(
            """
            INSERT INTO inventory_item (display_name, category, base_uom)
            VALUES (%s, %s, %s)
            RETURNING id
            """,
            (display_name, CATEGORY, BASE_UOM),
        )
        item_id = cur.fetchone()[0]
    conn.commit()
    return item_id


def db_repair_material(conn, item_id: int) -> None:
    """Fill any null required fields on an existing record."""
    with conn.cursor() as cur:
        cur.execute(
            """
            UPDATE inventory_item
            SET  category = COALESCE(category, %s),
                 base_uom = COALESCE(base_uom, %s),
                 updated_at = NOW()
            WHERE id = %s
            """,
            (CATEGORY, BASE_UOM, item_id),
        )
    conn.commit()


# ------------------------------------------------------------------ #
# Per-tab processing                                                   #
# ------------------------------------------------------------------ #

def process_tab(ws, tab_name: str, conn, logger: SeedLogger) -> str:
    """
    Process one workbook tab.
    Returns 'success' | 'skip' | 'fail'.
    """
    # Locate header row
    header_idx = find_header_row_idx(ws)
    if header_idx < 0:
        msg = "Header row not found — no row with both 'Items' and 'Date' columns"
        logger.log_failure(tab_name, msg, f"sheet={tab_name}")
        print(f"  FAIL  [{tab_name}] {msg}")
        return "fail"

    # Find first data row
    data_row, row_num = get_first_data_row(ws, header_idx)
    if data_row is None:
        msg = "No data rows — Items column is empty throughout"
        logger.log_failure(tab_name, msg, f"sheet={tab_name}, after row {header_idx + 1}")
        print(f"  FAIL  [{tab_name}] {msg}")
        return "fail"

    # Read Items value
    items_val = data_row[COL_ITEMS] if len(data_row) > COL_ITEMS else None
    if items_val is None or not str(items_val).strip():
        msg = "Items column is empty on first data row"
        logger.log_failure(tab_name, msg, f"sheet={tab_name}, row={row_num}, col=B")
        print(f"  FAIL  [{tab_name}] {msg}")
        return "fail"

    display_name = str(items_val).strip() + " MDF"

    # Check DB state
    try:
        item_id, is_complete = db_check_material(conn, display_name)
    except Exception as exc:
        conn.rollback()
        msg = f"DB query error: {exc}"
        logger.log_failure(tab_name, msg, f"display_name='{display_name}'")
        print(f"  FAIL  [{tab_name}] {msg}")
        return "fail"

    if item_id is not None and is_complete:
        print(f"  SKIP  [{tab_name}] '{display_name}' already complete in DB")
        logger.log_success(tab_name)
        return "skip"

    # Write to DB
    try:
        if item_id is not None:
            # Partial record — repair
            db_repair_material(conn, item_id)
            print(f"  REPAIR[{tab_name}] patched incomplete record '{display_name}' (id={item_id})")
        else:
            item_id = db_insert_material(conn, display_name)
            print(f"  INSERT[{tab_name}] '{display_name}' (id={item_id})")
        logger.log_success(tab_name)
        return "success"
    except Exception as exc:
        conn.rollback()
        # Unique-constraint violation means another tab already claimed this display_name.
        msg = f"DB write error: {exc}"
        detail = (
            f"display_name='{display_name}', source Items='{items_val}', "
            f"sheet='{tab_name}', row={row_num}. "
            "Possible Items-column data entry error: another tab may use the same Items value."
        )
        logger.log_failure(tab_name, msg, detail)
        print(f"  FAIL  [{tab_name}] {msg}")
        print(f"        {detail}")
        return "fail"


# ------------------------------------------------------------------ #
# Entry point                                                          #
# ------------------------------------------------------------------ #

def main():
    parser = argparse.ArgumentParser(
        description="DE-1.1: Insert board materials from Stock_management.xlsx"
    )
    parser.add_argument(
        "--retry",
        action="store_true",
        help="Process only tabs listed in seed_failures.log (default: auto-detect)",
    )
    args = parser.parse_args()

    logger = SeedLogger(SCRIPT_NAME)

    # Determine run mode
    failed_keys = logger.get_failed_keys()
    is_retry = args.retry or bool(failed_keys)

    if is_retry:
        print(f"RETRY mode — {len(failed_keys)} tab(s) pending: {sorted(failed_keys)}\n")
    else:
        print("FULL mode — scanning all valid tabs\n")

    # Open workbook
    try:
        wb = openpyxl.load_workbook(WORKBOOK_PATH, data_only=True)
    except Exception as exc:
        print(f"ERROR: Cannot open workbook: {exc}", file=sys.stderr)
        sys.exit(1)

    # Connect to DB
    conn = get_connection()

    succeeded = failed = skipped = 0

    for tab_name in wb.sheetnames:
        if not is_valid_tab(tab_name):
            print(f"  IGNORE[{tab_name}] — not a thickness tab")
            continue

        if is_retry and tab_name not in failed_keys:
            print(f"  SKIP  [{tab_name}] — not in retry list (previously succeeded)")
            skipped += 1
            continue

        ws = wb[tab_name]
        result = process_tab(ws, tab_name, conn, logger)

        if result == "success":
            succeeded += 1
        elif result == "skip":
            skipped += 1
        else:
            failed += 1

    conn.close()

    print("\n" + "=" * 52)
    print("DE-1.1 SUMMARY")
    print(f"  Succeeded : {succeeded}")
    print(f"  Skipped   : {skipped}  (already complete)")
    print(f"  Failed    : {failed}")
    if failed:
        from seed.seed_logger import FAILURES_LOG
        print(f"  Failure log : {FAILURES_LOG}")
    print("=" * 52)

    sys.exit(1 if failed else 0)


if __name__ == "__main__":
    main()
