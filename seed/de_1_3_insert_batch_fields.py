"""
DE-1.3: Read supplier workbook headers and insert per-batch field definitions.

Reads the header row from the 'Received tree logs trucks' sheet in
Tree_log_suppliers_Book.xlsx, maps each column name to a batch_field
record, and inserts it against every board material in the DB.

Failure handling:
  - Missing expected header → logged with expected vs actual list; does not stop.
  - Material missing from DB → skipped (logged as blocked on DE-1.1).
  - Field exists with different is_required flag → logged as conflict; NOT overwritten.
  - Field exists and identical → skipped (idempotent).

Idempotency key for retry: "<display_name>::<field_key>"
  e.g. "6mm MDF::date", "9mm MDF::plate_number"

Usage:
    python -m seed.de_1_3_insert_batch_fields
    python -m seed.de_1_3_insert_batch_fields --retry
"""

import argparse
import os
import sys

import openpyxl

import json
from datetime import datetime, timezone

from config import LOADED_DIR, RAW_DATA_DIR, ROOT, TRACKER_PATH
from etl.file_tracker import FileTracker
from seed.db import get_connection
from seed.seed_logger import SeedLogger

SCRIPT_NAME = "de_1_3"
SUPPLIER_WORKBOOK = str(RAW_DATA_DIR / "Tree log suppliers Book.xlsx")
SUPPLIER_SHEET = "Received tree logs trucks"

# Canonical mapping: workbook column header → batch_field attributes.
# Strip whitespace from actual headers before matching.
COLUMN_MAPPING = {
    "Date":                    {"field_key": "date",              "field_type": "date",    "is_required": True},
    "Supplier name":           {"field_key": "supplier_name",     "field_type": "text",    "is_required": True},
    "Plate number":            {"field_key": "plate_number",      "field_type": "text",    "is_required": True},
    "Log length (m)":          {"field_key": "log_length_m",      "field_type": "decimal", "is_required": True},
    "Supplied Wood width (m)": {"field_key": "wood_width_m",      "field_type": "decimal", "is_required": True},
    "Supplied Wood length (m)":{"field_key": "wood_length_m",     "field_type": "decimal", "is_required": True},
    "Total cubed m3":          {"field_key": "total_m3",          "field_type": "decimal", "is_required": False},
    "Unit price per m3":       {"field_key": "unit_price_per_m3", "field_type": "decimal", "is_required": True},
    "Total price":             {"field_key": "total_price",       "field_type": "decimal", "is_required": False},
    "Unqualified (Rwf)":       {"field_key": "unqualified_rwf",   "field_type": "decimal", "is_required": False},
    "Return out":              {"field_key": "return_out",        "field_type": "decimal", "is_required": False},
    "Balance to be paid":      {"field_key": "balance_to_be_paid","field_type": "decimal", "is_required": False},
    "Paid amount":             {"field_key": "paid_amount",       "field_type": "decimal", "is_required": False},
    "Payment date":            {"field_key": "payment_date",      "field_type": "date",    "is_required": False},
    "Payment Mode":            {"field_key": "payment_mode",      "field_type": "text",    "is_required": False},
    "Cheque number":           {"field_key": "cheque_number",     "field_type": "text",    "is_required": False},
    "Remaining balance":       {"field_key": "remaining_balance", "field_type": "decimal", "is_required": False},
}

EXPECTED_HEADERS = set(COLUMN_MAPPING.keys())


# ------------------------------------------------------------------ #
# Workbook helpers                                                     #
# ------------------------------------------------------------------ #

def read_header_row(wb, sheet_name: str, logger: SeedLogger):
    """
    Read and return the list of non-null, stripped header names from the sheet.
    Logs any expected header that is absent.
    Returns (field_defs, actual_headers) or (None, actual_headers) on fatal error.
    """
    if sheet_name not in wb.sheetnames:
        logger.log_failure(
            sheet_name,
            f"Sheet '{sheet_name}' not found in workbook",
            f"Available sheets: {wb.sheetnames}",
        )
        return None, []

    ws = wb[sheet_name]
    # Header row is the first row that contains >= 2 non-null cells
    actual_headers = []
    for row in ws.iter_rows(values_only=True):
        cells = [str(c).strip() for c in row if c is not None and str(c).strip()]
        if len(cells) >= 2:
            actual_headers = cells
            break

    if not actual_headers:
        logger.log_failure(
            sheet_name,
            "Could not locate header row (no row with >= 2 non-null cells)",
            f"sheet={sheet_name}",
        )
        return None, []

    # Check for missing expected headers
    actual_set = set(actual_headers)
    for expected in EXPECTED_HEADERS:
        if expected not in actual_set:
            logger.log_failure(
                f"header::{expected}",
                f"Expected header '{expected}' not found in '{sheet_name}'",
                f"Expected: {sorted(EXPECTED_HEADERS)}\nActual: {actual_headers}",
            )
            print(f"  WARN  Missing header '{expected}' — logged for diagnosis")

    # Build field_defs list from headers that are in the mapping
    field_defs = []
    for header in actual_headers:
        if header in COLUMN_MAPPING:
            field_defs.append({"field_name": header, **COLUMN_MAPPING[header]})

    return field_defs, actual_headers


# ------------------------------------------------------------------ #
# Database helpers                                                     #
# ------------------------------------------------------------------ #

def db_get_all_materials(conn) -> list:
    with conn.cursor() as cur:
        cur.execute("SELECT id, display_name FROM inventory_item ORDER BY display_name")
        return [{"id": r[0], "display_name": r[1]} for r in cur.fetchall()]


def db_check_field(conn, item_id: int, field_key: str):
    """
    Returns (exists: bool, conflicts: bool, existing_required: bool|None).
    conflicts is True when the field exists but is_required differs.
    """
    with conn.cursor() as cur:
        cur.execute(
            "SELECT is_required FROM batch_field WHERE item_id = %s AND field_key = %s",
            (item_id, field_key),
        )
        row = cur.fetchone()
    if row is None:
        return False, False, None
    existing_required = row[0]
    return True, None, existing_required  # conflict check done by caller


def db_insert_field(conn, item_id: int, field_name: str, field_def: dict) -> None:
    with conn.cursor() as cur:
        cur.execute(
            """
            INSERT INTO batch_field (item_id, field_name, field_key, field_type, is_required)
            VALUES (%s, %s, %s, %s, %s)
            """,
            (
                item_id,
                field_name,
                field_def["field_key"],
                field_def["field_type"],
                field_def["is_required"],
            ),
        )
    conn.commit()


# ------------------------------------------------------------------ #
# Per-material processing                                             #
# ------------------------------------------------------------------ #

def process_material(
    material: dict,
    field_defs: list,
    conn,
    logger: SeedLogger,
    retry_keys: set,
) -> dict:
    item_id = material["id"]
    display_name = material["display_name"]

    stats = {
        "display_name": display_name,
        "inserted": 0,
        "skipped": 0,
        "blocked": 0,
        "conflicts": 0,
    }

    for fd in field_defs:
        field_key = fd["field_key"]
        retry_key = f"{display_name}::{field_key}"

        if retry_keys and retry_key not in retry_keys:
            continue

        exists, _, existing_required = db_check_field(conn, item_id, field_key)

        if exists:
            if existing_required != fd["is_required"]:
                # Conflict: different required flag — do NOT overwrite automatically
                logger.log_failure(
                    retry_key,
                    (
                        f"Conflict: field '{field_key}' on '{display_name}' exists with "
                        f"is_required={existing_required} but mapping expects is_required={fd['is_required']}. "
                        "Manual review required before next run."
                    ),
                    f"item_id={item_id}, field_key={field_key}",
                )
                print(
                    f"  CONFLICT [{display_name}] field '{field_key}': "
                    f"existing is_required={existing_required} vs expected {fd['is_required']}"
                )
                stats["conflicts"] += 1
            else:
                logger.log_success(retry_key)
                stats["skipped"] += 1
            continue

        try:
            db_insert_field(conn, item_id, fd["field_name"], fd)
            logger.log_success(retry_key)
            print(f"  INSERT [{display_name}] field '{field_key}' ({fd['field_type']}, required={fd['is_required']})")
            stats["inserted"] += 1
        except Exception as exc:
            conn.rollback()
            logger.log_failure(
                retry_key,
                f"DB insert error for field '{field_key}': {exc}",
                f"item_id={item_id}, field_key={field_key}",
            )
            print(f"  FAIL  [{display_name}] field '{field_key}': {exc}")

    return stats


# ------------------------------------------------------------------ #
# Entry point                                                          #
# ------------------------------------------------------------------ #

def main():
    parser = argparse.ArgumentParser(description="DE-1.3: Insert batch field definitions from supplier workbook")
    parser.add_argument("--retry", action="store_true")
    args = parser.parse_args()

    logger = SeedLogger(SCRIPT_NAME)
    failed_keys = logger.get_failed_keys()
    is_retry = args.retry or bool(failed_keys)

    if is_retry:
        print(f"RETRY mode — {len(failed_keys)} key(s) pending\n")
    else:
        print("FULL mode — processing all materials\n")

    try:
        wb = openpyxl.load_workbook(SUPPLIER_WORKBOOK, data_only=True, read_only=True)
    except Exception as exc:
        print(f"ERROR: Cannot open supplier workbook: {exc}", file=sys.stderr)
        sys.exit(1)

    print(f"Reading headers from '{SUPPLIER_SHEET}' ...")
    field_defs, actual_headers = read_header_row(wb, SUPPLIER_SHEET, logger)
    if field_defs is None:
        print("ERROR: Cannot read supplier sheet headers — see seed_failures.log", file=sys.stderr)
        sys.exit(1)

    print(f"  Found {len(field_defs)} mapped field(s) from {len(actual_headers)} header(s)\n")

    conn = get_connection()
    materials = db_get_all_materials(conn)

    if not materials:
        print("WARNING: No materials found in DB — run DE-1.1 first.")
        conn.close()
        sys.exit(0)

    all_stats = []
    total_conflicts = 0

    for material in materials:
        display_name = material["display_name"]

        if is_retry:
            mat_keys = {k for k in failed_keys if k.startswith(f"{display_name}::")}
            if not mat_keys:
                print(f"  SKIP  [{display_name}] — no pending field keys")
                continue
            retry_keys = mat_keys
        else:
            retry_keys = set()

        stats = process_material(material, field_defs, conn, logger, retry_keys)
        all_stats.append(stats)
        total_conflicts += stats["conflicts"]

    conn.close()

    print("\n" + "=" * 60)
    print("DE-1.3 SUMMARY (per material)")
    for s in all_stats:
        print(
            f"  {s['display_name']:<18} inserted={s['inserted']}  "
            f"skipped={s['skipped']}  blocked={s['blocked']}  conflicts={s['conflicts']}"
        )

    remaining = logger.get_failed_keys()
    print(f"\n  Pending failures: {len(remaining)}")
    if total_conflicts:
        print(f"  Conflicts requiring manual review: {total_conflicts}")
        print("  Review seed_failures.log and resolve before re-running.")
    if remaining:
        from seed.seed_logger import FAILURES_LOG
        print(f"  Failure log: {FAILURES_LOG}")
    print("=" * 60)

    if not remaining:
        source_path = RAW_DATA_DIR / "Tree log suppliers Book.xlsx"
        LOADED_DIR.mkdir(parents=True, exist_ok=True)
        total_inserted = sum(s["inserted"] for s in all_stats)
        manifest = {
            "script": SCRIPT_NAME,
            "source_file": str(source_path.relative_to(ROOT)),
            "mode": "retry" if is_retry else "full",
            "loaded_at": datetime.now(timezone.utc).isoformat(),
            "summary": {
                "materials_processed": len(all_stats),
                "fields_inserted": total_inserted,
                "conflicts": total_conflicts,
                "failed": len(remaining),
            },
        }
        with open(LOADED_DIR / "tree_log_suppliers_de_1_3.json", "w", encoding="utf-8") as f:
            json.dump(manifest, f, indent=2)
        FileTracker(TRACKER_PATH, ROOT).mark_loaded(source_path)
        print(f"  Load manifest: {LOADED_DIR / 'tree_log_suppliers_de_1_3.json'}")

    sys.exit(1 if remaining else 0)


if __name__ == "__main__":
    main()
