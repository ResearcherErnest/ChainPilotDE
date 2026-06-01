"""
DE-1.5: Insert adhesive/glue materials from all glue_ledger sources.

Reads every source in config.SOURCES with seed_type="glue_ledger",
extracts the Items column value from the first data row of each file,
and inserts one inventory_item per source:
    display_name = <Items value, title-cased>   (e.g. "Glue")
    category     = 'Raw Material'
    base_uom     = source["base_uom"]           (e.g. "bags")

To add a new adhesive source, append one entry to SOURCES in config.py
with seed_type="glue_ledger" — no changes needed here.

Idempotency:
  - Record already complete → skip.
  - Partial record → repair missing fields.

Usage:
    python -m seed.de_1_5_insert_glue
    python -m seed.de_1_5_insert_glue --retry
"""

import argparse
import json
import sys
from datetime import datetime, timezone

import openpyxl

from config import LOADED_DIR, RAW_DATA_DIR, ROOT, SOURCES, TRACKER_PATH
from etl.file_tracker import FileTracker
from seed.db import get_connection
from seed.seed_logger import SeedLogger

SCRIPT_NAME = "de_1_5"
CATEGORY    = "Raw Material"
COL_ITEMS   = 1   # column B (0-based)

GLUE_SOURCES = [s for s in SOURCES if s.get("seed_type") == "glue_ledger"]


# ------------------------------------------------------------------ #
# Workbook helpers                                                     #
# ------------------------------------------------------------------ #

def find_header_row_idx(ws) -> int:
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        vals = [str(v).strip().lower() for v in row if v is not None]
        if "items" in vals and "date" in vals:
            return i
    return -1


def get_display_name(ws, header_idx: int):
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i <= header_idx:
            continue
        val = row[COL_ITEMS] if len(row) > COL_ITEMS else None
        if val is not None and str(val).strip():
            return str(val).strip().title()
    return None


# ------------------------------------------------------------------ #
# Database helpers                                                     #
# ------------------------------------------------------------------ #

def db_check(conn, display_name: str):
    with conn.cursor() as cur:
        cur.execute(
            "SELECT id, category, base_uom FROM inventory_item WHERE display_name = %s",
            (display_name,),
        )
        row = cur.fetchone()
    if row is None:
        return None, False
    item_id, category, base_uom = row
    return item_id, (category is not None and base_uom is not None)


def db_insert(conn, display_name: str, base_uom: str) -> int:
    with conn.cursor() as cur:
        cur.execute(
            "INSERT INTO inventory_item (display_name, category, base_uom) VALUES (%s, %s, %s) RETURNING id",
            (display_name, CATEGORY, base_uom),
        )
        item_id = cur.fetchone()[0]
    conn.commit()
    return item_id


def db_repair(conn, item_id: int, base_uom: str):
    with conn.cursor() as cur:
        cur.execute(
            """UPDATE inventory_item
               SET category = COALESCE(category, %s),
                   base_uom = COALESCE(base_uom, %s),
                   updated_at = NOW()
               WHERE id = %s""",
            (CATEGORY, base_uom, item_id),
        )
    conn.commit()


# ------------------------------------------------------------------ #
# Per-source processing                                                #
# ------------------------------------------------------------------ #

def process_source(source: dict, conn, logger: SeedLogger, is_retry: bool, failed_keys: set) -> str:
    """
    Process one glue_ledger source. Returns 'success' | 'skip' | 'fail'.
    """
    filepath = RAW_DATA_DIR / source["filename"]
    base_uom = source.get("base_uom", "units")
    retry_key = source["source_id"]

    if is_retry and retry_key not in failed_keys:
        print(f"  SKIP  [{source['filename']}] — not in retry list (previously succeeded)")
        return "skip"

    try:
        wb = openpyxl.load_workbook(str(filepath), data_only=True, read_only=True)
    except Exception as exc:
        msg = f"Cannot open workbook: {exc}"
        logger.log_failure(retry_key, msg, str(filepath))
        print(f"  FAIL  [{source['filename']}] {msg}")
        return "fail"

    ws = wb.worksheets[0]
    header_idx = find_header_row_idx(ws)
    if header_idx < 0:
        wb.close()
        msg = "Header row not found"
        logger.log_failure(retry_key, msg, f"sheet={ws.title}")
        print(f"  FAIL  [{source['filename']}] {msg}")
        return "fail"

    display_name = get_display_name(ws, header_idx)
    wb.close()

    if display_name is None:
        msg = "Items column empty — cannot determine material name"
        logger.log_failure(retry_key, msg, f"file={source['filename']}")
        print(f"  FAIL  [{source['filename']}] {msg}")
        return "fail"

    try:
        item_id, is_complete = db_check(conn, display_name)
    except Exception as exc:
        conn.rollback()
        msg = f"DB query error: {exc}"
        logger.log_failure(retry_key, msg, f"display_name='{display_name}'")
        print(f"  FAIL  [{source['filename']}] {msg}")
        return "fail"

    if item_id is not None and is_complete:
        logger.log_success(retry_key)
        print(f"  SKIP  [{source['filename']}] '{display_name}' already complete (id={item_id})")
        return "skip"

    try:
        if item_id is not None:
            db_repair(conn, item_id, base_uom)
            print(f"  REPAIR[{source['filename']}] patched '{display_name}' (id={item_id})")
        else:
            item_id = db_insert(conn, display_name, base_uom)
            print(f"  INSERT[{source['filename']}] '{display_name}' base_uom='{base_uom}' (id={item_id})")
        logger.log_success(retry_key)
        return "success"
    except Exception as exc:
        conn.rollback()
        msg = f"DB write error: {exc}"
        logger.log_failure(retry_key, msg, f"display_name='{display_name}'")
        print(f"  FAIL  [{source['filename']}] {msg}")
        return "fail"


# ------------------------------------------------------------------ #
# Entry point                                                          #
# ------------------------------------------------------------------ #

def main():
    parser = argparse.ArgumentParser(description="DE-1.5: Insert glue/adhesive materials from all glue_ledger sources")
    parser.add_argument("--retry", action="store_true")
    args = parser.parse_args()

    if not GLUE_SOURCES:
        print("No glue_ledger sources configured in config.SOURCES — nothing to do.")
        sys.exit(0)

    logger = SeedLogger(SCRIPT_NAME)
    failed_keys = logger.get_failed_keys()
    is_retry = args.retry or bool(failed_keys)

    print(f"{'RETRY' if is_retry else 'FULL'} mode — {len(GLUE_SOURCES)} glue_ledger source(s)\n")

    conn = get_connection()
    tracker = FileTracker(TRACKER_PATH, ROOT)

    succeeded = skipped = failed = 0

    for source in GLUE_SOURCES:
        result = process_source(source, conn, logger, is_retry, failed_keys)
        if result == "success":
            succeeded += 1
            tracker.mark_loaded(RAW_DATA_DIR / source["filename"])
        elif result == "skip":
            skipped += 1
        else:
            failed += 1

    conn.close()

    print("\n" + "=" * 52)
    print("DE-1.5 SUMMARY")
    print(f"  Sources   : {len(GLUE_SOURCES)}")
    print(f"  Succeeded : {succeeded}")
    print(f"  Skipped   : {skipped}  (already complete)")
    print(f"  Failed    : {failed}")
    print("=" * 52)

    if not failed:
        LOADED_DIR.mkdir(parents=True, exist_ok=True)
        manifest = {
            "script": SCRIPT_NAME,
            "sources_processed": [s["source_id"] for s in GLUE_SOURCES],
            "mode": "retry" if is_retry else "full",
            "loaded_at": datetime.now(timezone.utc).isoformat(),
            "summary": {"succeeded": succeeded, "skipped": skipped, "failed": failed},
        }
        with open(LOADED_DIR / "glue_materials_de_1_5.json", "w", encoding="utf-8") as f:
            json.dump(manifest, f, indent=2)
        print(f"  Load manifest: {LOADED_DIR / 'glue_materials_de_1_5.json'}")

    sys.exit(1 if failed else 0)


if __name__ == "__main__":
    main()
