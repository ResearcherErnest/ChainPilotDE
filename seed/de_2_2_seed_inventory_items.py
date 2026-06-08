"""
DE-2.2: Seed inventory items into the backend schema.

Target table:
  inventory_item — one row per material with attributes and batch_fields stored as JSONB

  attributes JSONB — item specifications:
    [{"label": "Thickness", "value": "0.9"},
     {"label": "Width",     "value": "122.0"},
     {"label": "Length",    "value": "244.0"}]

  batch_fields JSONB — intake form field definitions:
    [{"code": "date", "label": "Date", "data_type": "date", "is_required": true, "sort_order": 1}, ...]

Prerequisites:
    python -m seed.de_2_1_seed_reference_data

Idempotency:
  - Item with same SKU already in inventory_item → skip.

Usage:
    python -m seed.de_2_2_seed_inventory_items
    python -m seed.de_2_2_seed_inventory_items --dry-run
    python -m seed.de_2_2_seed_inventory_items --retry
"""

import argparse
import json
import re
import sys
from datetime import datetime, timezone
from decimal import Decimal, InvalidOperation

import openpyxl

from config import LOADED_DIR, RAW_DATA_DIR, ROOT, SOURCES, TRACKER_PATH
from etl.file_tracker import FileTracker
from seed.db import get_connection
from seed.schema_registry import (
    GLUE_ATTRIBUTE_LABELS,
    MDF_ATTRIBUTE_LABELS,
    MDF_BATCH_FIELDS,
    RAW_MATERIAL_CODE,
    UOM_CODE_MAP,
)
from seed.seed_logger import SeedLogger

SCRIPT_NAME  = "de_2_2"
THICKNESS_RE = re.compile(r"^\d+\s*[Mm]{2}\s*$")
GLUE_SPEC_RE = re.compile(r"^(\d+(?:\.\d+)?)kg/(\d+)bag$", re.IGNORECASE)
COL_ITEMS    = 1
COL_SPECS    = 2


# ── Reference-data lookups ────────────────────────────────────────────────

def fetch_category_id(conn, code: str) -> str:
    with conn.cursor() as cur:
        cur.execute("SELECT id FROM category WHERE code = %s", (code,))
        row = cur.fetchone()
    if not row:
        raise RuntimeError(
            f"Category '{code}' not found — run de_2_1_seed_reference_data first"
        )
    return str(row[0])


def fetch_uom_id(conn, code: str) -> str:
    with conn.cursor() as cur:
        cur.execute("SELECT id FROM unit_of_measure WHERE code = %s", (code,))
        row = cur.fetchone()
    if not row:
        raise RuntimeError(
            f"UOM '{code}' not found — run de_2_1_seed_reference_data first"
        )
    return str(row[0])


# ── Workbook helpers ──────────────────────────────────────────────────────

def find_header_row(ws, required_lower: list) -> int:
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        vals = {str(v).strip().lower() for v in row if v is not None}
        if all(c in vals for c in required_lower):
            return i
    return -1


def first_cell_value(ws, header_idx: int, col: int):
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i <= header_idx:
            continue
        val = row[col] if len(row) > col else None
        if val is not None and str(val).strip():
            return str(val).strip()
    return None


# ── Spec parsers ──────────────────────────────────────────────────────────

def parse_mdf_spec(raw: str):
    parts = raw.split("*")
    if len(parts) != 3:
        return [], f"Expected T*W*L, got: '{raw}'"
    try:
        values = [str(Decimal(p.strip())) for p in parts]
    except InvalidOperation as exc:
        return [], f"Non-numeric segment in '{raw}': {exc}"
    return [{"label": lbl, "value": v} for lbl, v in zip(MDF_ATTRIBUTE_LABELS, values)], None


def parse_glue_spec(raw: str):
    m = GLUE_SPEC_RE.match(raw.strip())
    if not m:
        return [], f"Unrecognised glue spec '{raw}' — expected e.g. '25kg/1bag'"
    try:
        values = [str(Decimal(m.group(1))), str(Decimal(m.group(2)))]
    except InvalidOperation as exc:
        return [], f"Non-numeric in '{raw}': {exc}"
    return [{"label": lbl, "value": v} for lbl, v in zip(GLUE_ATTRIBUTE_LABELS, values)], None


# ── Database helpers ──────────────────────────────────────────────────────

def item_id_by_sku(conn, sku: str):
    with conn.cursor() as cur:
        cur.execute("SELECT id FROM inventory_item WHERE sku = %s", (sku,))
        row = cur.fetchone()
    return str(row[0]) if row else None


def insert_item(conn, sku, name, category_id, uom_id, attributes, batch_fields) -> str:
    with conn.cursor() as cur:
        cur.execute(
            """
            INSERT INTO inventory_item (sku, name, category_id, uom_id, attributes, batch_fields)
            VALUES (%s, %s, %s::uuid, %s::uuid, %s::jsonb, %s::jsonb)
            RETURNING id
            """,
            (sku, name, category_id, uom_id,
             json.dumps(attributes) if attributes else None,
             json.dumps(batch_fields) if batch_fields else None),
        )
        item_id = str(cur.fetchone()[0])
    conn.commit()
    return item_id


# ── Per-item seeding ──────────────────────────────────────────────────────

def seed_item(conn, sku, name, category_id, uom_id, attributes,
              batch_fields, logger, dry_run) -> str:
    existing = item_id_by_sku(conn, sku)
    if existing:
        logger.log_success(sku)
        detail = f"name='{name}'  id={existing}"
        logger.log_event("SKIP", f"inventory_item:{sku}", detail)
        print(f"  SKIP   [{sku}] '{name}' already exists (id={existing})")
        return "skip"

    if dry_run:
        attrs = ", ".join(f"{a['label']}={a['value']}" for a in attributes)
        print(f"  DRY    [{sku}] '{name}'  [{attrs}]  batch_fields={len(batch_fields)}")
        return "skip"

    try:
        item_id = insert_item(conn, sku, name, category_id, uom_id, attributes, batch_fields)
        logger.log_success(sku)
        attr_summary = "  ".join(f"{a['label']}={a['value']}" for a in attributes)
        logger.log_event(
            "CREATE", f"inventory_item:{sku}",
            f"name='{name}'  id={item_id}  attrs=[{attr_summary}]  batch_fields={len(batch_fields)}",
        )
        print(f"  CREATE [{sku}] '{name}' id={item_id}  batch_fields={len(batch_fields)}")
        return "created"
    except Exception as exc:
        conn.rollback()
        logger.log_failure(sku, f"DB error: {exc}", f"name='{name}'")
        logger.log_event("FAIL", f"inventory_item:{sku}", f"name='{name}'  error={exc}")
        print(f"  FAIL   [{sku}] {exc}")
        return "fail"


# ── MDF processing ────────────────────────────────────────────────────────

def process_mdf(conn, logger, category_id, uom_id,
                failed_keys, is_retry, dry_run):
    src = next((s for s in SOURCES if s.get("seed_type") == "mdf_ledger"), None)
    if not src:
        print("  No mdf_ledger source configured — skipping")
        return 0, 0, 0

    try:
        wb = openpyxl.load_workbook(
            str(RAW_DATA_DIR / src["filename"]), data_only=True, read_only=True
        )
    except Exception as exc:
        print(f"ERROR opening {src['filename']}: {exc}", file=sys.stderr)
        sys.exit(1)

    created = skipped = failed = 0
    for tab in wb.sheetnames:
        if not THICKNESS_RE.match(tab):
            continue

        sku = f"RM-MDF-{tab.strip().upper()}"
        if is_retry and sku not in failed_keys:
            skipped += 1
            continue

        ws  = wb[tab]
        hdr = find_header_row(ws, ["items", "specifications"])
        if hdr < 0:
            logger.log_failure(sku, "Header row not found", f"sheet={tab}")
            print(f"  FAIL   [{sku}] no header in tab '{tab}'")
            failed += 1
            continue

        name     = tab.strip().upper() + " MDF Board"
        spec_raw = first_cell_value(ws, hdr, COL_SPECS)
        attrs, err = parse_mdf_spec(spec_raw) if spec_raw else ([], None)
        if err:
            print(f"  WARN   [{sku}] spec parse: {err} — attributes omitted")

        result = seed_item(conn, sku, name, category_id, uom_id,
                           attrs, MDF_BATCH_FIELDS, logger, dry_run)
        if result == "created":   created += 1
        elif result == "skip":    skipped += 1
        else:                     failed  += 1

    wb.close()
    return created, skipped, failed


# ── Glue processing ───────────────────────────────────────────────────────

def process_glue(conn, logger, category_id, uom_ids,
                 failed_keys, is_retry, dry_run):
    sources = [s for s in SOURCES if s.get("seed_type") == "glue_ledger"]
    created = skipped = failed = 0

    for idx, src in enumerate(sources, start=1):
        sku      = f"RM-GLUE-{idx:03d}"
        uom_code = UOM_CODE_MAP.get(src.get("base_uom", ""), "PIECE")
        uom_id   = uom_ids.get(uom_code)
        if not uom_id:
            logger.log_failure(sku, f"UOM '{uom_code}' not found", "")
            print(f"  FAIL   [{sku}] UOM '{uom_code}' missing — run de_2_1 first")
            failed += 1
            continue

        if is_retry and sku not in failed_keys:
            skipped += 1
            continue

        try:
            wb = openpyxl.load_workbook(
                str(RAW_DATA_DIR / src["filename"]), data_only=True, read_only=True
            )
        except Exception as exc:
            logger.log_failure(sku, f"Cannot open workbook: {exc}", src["filename"])
            print(f"  FAIL   [{sku}] {exc}")
            failed += 1
            continue

        ws  = wb.worksheets[0]
        hdr = find_header_row(ws, ["items", "specifications"])
        if hdr < 0:
            wb.close()
            logger.log_failure(sku, "Header row not found", src["filename"])
            print(f"  FAIL   [{sku}] no header in '{src['filename']}'")
            failed += 1
            continue

        items_val = first_cell_value(ws, hdr, COL_ITEMS)
        spec_raw  = first_cell_value(ws, hdr, COL_SPECS)
        wb.close()

        if not items_val:
            logger.log_failure(sku, "Items column empty", src["filename"])
            print(f"  FAIL   [{sku}] Items column empty")
            failed += 1
            continue

        name       = items_val.title()
        attrs, err = parse_glue_spec(spec_raw) if spec_raw else ([], None)
        if err:
            print(f"  WARN   [{sku}] spec parse: {err} — attributes omitted")

        result = seed_item(conn, sku, name, category_id, uom_id,
                           attrs, [], logger, dry_run)
        if result == "created":   created += 1
        elif result == "skip":    skipped += 1
        else:                     failed  += 1

    return created, skipped, failed


# ── Entry point ───────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(
        description="DE-2.2: Seed inventory items into inventory_item (attributes + batch_fields as JSONB)"
    )
    parser.add_argument("--dry-run", action="store_true",
                        help="Preview without writing to DB")
    parser.add_argument("--retry", action="store_true",
                        help="Process only SKUs listed in seed_failures.log")
    args = parser.parse_args()

    logger      = SeedLogger(SCRIPT_NAME)
    failed_keys = logger.get_failed_keys()
    is_retry    = args.retry or bool(failed_keys)

    mode = "retry" if is_retry else "dry-run" if args.dry_run else "full"
    print(f"{mode.upper()} mode\n")
    logger.log_run_start(
        description="Seed inventory_item: attributes + batch_fields JSONB from Excel sources",
        mode=mode,
    )

    conn = get_connection()

    try:
        category_id  = fetch_category_id(conn, RAW_MATERIAL_CODE)
        sheet_uom_id = fetch_uom_id(conn, "SHEET")
        piece_uom_id = fetch_uom_id(conn, "PIECE")
    except RuntimeError as exc:
        print(f"ERROR: {exc}", file=sys.stderr)
        conn.close()
        sys.exit(1)

    uom_ids = {"SHEET": sheet_uom_id, "PIECE": piece_uom_id}
    print(f"  Raw Material id : {category_id}")
    print(f"  SHEET uom id    : {sheet_uom_id}")
    print(f"  PIECE uom id    : {piece_uom_id}\n")

    print("-" * 56)
    print("MDF Boards")
    print("-" * 56)
    mdf_c, mdf_s, mdf_f = process_mdf(
        conn, logger, category_id, sheet_uom_id,
        failed_keys, is_retry, args.dry_run,
    )

    print("\n" + "-" * 56)
    print("Glue / Adhesives")
    print("-" * 56)
    glue_c, glue_s, glue_f = process_glue(
        conn, logger, category_id, uom_ids,
        failed_keys, is_retry, args.dry_run,
    )

    conn.close()

    remaining = logger.get_failed_keys()
    print("\n" + "=" * 56)
    print("DE-2.2 SUMMARY")
    print(f"  MDF boards  : created={mdf_c}  skipped={mdf_s}  failed={mdf_f}")
    print(f"  Glue        : created={glue_c}  skipped={glue_s}  failed={glue_f}")
    print(f"  Pending failures : {len(remaining)}")
    print("=" * 56)

    if not remaining and not args.dry_run:
        LOADED_DIR.mkdir(parents=True, exist_ok=True)
        manifest = {
            "script":    SCRIPT_NAME,
            "mode":      "retry" if is_retry else "full",
            "loaded_at": datetime.now(timezone.utc).isoformat(),
            "summary":   {"mdf_created": mdf_c, "glue_created": glue_c,
                          "total_failed": mdf_f + glue_f},
        }
        manifest_path = LOADED_DIR / "inventory_items_de_2_2.json"
        with open(manifest_path, "w", encoding="utf-8") as f:
            json.dump(manifest, f, indent=2)
        tracker = FileTracker(TRACKER_PATH, ROOT)
        for s in SOURCES:
            if s.get("seed_type") in ("mdf_ledger", "glue_ledger"):
                tracker.mark_loaded(RAW_DATA_DIR / s["filename"])
        print(f"  Manifest: {manifest_path}")

    run_summary = (f"MDF: created={mdf_c} skipped={mdf_s} failed={mdf_f}  "
                   f"Glue: created={glue_c} skipped={glue_s} failed={glue_f}  "
                   f"pending_failures={len(remaining)}")
    logger.log_run_end("error" if remaining else "ok", run_summary)

    sys.exit(1 if remaining else 0)


if __name__ == "__main__":
    main()
