"""
DE-1.2: Parse Specifications strings and insert dimension properties.

For each valid thickness tab in Stock_management.xlsx:
  1. Find the material in the DB (by deriving display_name from the Items column).
  2. Scan all rows in the Specifications column; collect distinct spec strings.
  3. Log informational warnings for variant specs (e.g. 6mm tab has three variants).
  4. Parse the canonical (first non-null) spec on '*' → thickness, width, length.
  5. Insert each property individually; on retry skip already-present properties.

Failure handling:
  - Material missing from DB → skipped (logged as blocked on DE-1.1).
  - Unparseable spec string → logged per-row in seed_failures.log; rest of tab continues.
  - Property insertion failure → logged; individual retry tracks per-property.

Idempotency key for retry: "<tab_name>.<property_name>"
  e.g. "6MM.thickness", "6MM.width", "6MM.length"

Usage:
    python -m seed.de_1_2_parse_specs
    python -m seed.de_1_2_parse_specs --retry
"""

import argparse
import os
import sys
from decimal import Decimal, InvalidOperation

import openpyxl

sys.path.insert(0, os.path.dirname(os.path.dirname(__file__)))
from seed.db import get_connection
from seed.seed_logger import SeedLogger

SCRIPT_NAME = "de_1_2"
WORKBOOK_PATH = os.path.join(
    os.path.dirname(os.path.dirname(__file__)), "Raw_data", "Stock management.xlsx"
)

import re
THICKNESS_RE = re.compile(r"^\d+\s*[Mm]{2}\s*$")

COL_ITEMS = 1
COL_SPECS = 2

PROPERTIES = ("thickness", "width", "length")


def is_valid_tab(name: str) -> bool:
    return bool(THICKNESS_RE.match(name))


def find_header_row_idx(ws) -> int:
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        vals = [str(v).strip().lower() for v in row if v is not None]
        if "items" in vals and "specifications" in vals:
            return i
    return -1


def get_display_name_from_tab(ws, header_idx: int):
    """Derive display_name the same way DE-1.1 does: Items column + ' MDF'."""
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i <= header_idx:
            continue
        items_val = row[COL_ITEMS] if len(row) > COL_ITEMS else None
        if items_val is not None and str(items_val).strip():
            return str(items_val).strip() + " MDF"
    return None


def collect_spec_strings(ws, header_idx: int) -> list:
    """Return list of (row_number, spec_string) for all non-null Specifications cells."""
    results = []
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i <= header_idx:
            continue
        spec_val = row[COL_SPECS] if len(row) > COL_SPECS else None
        if spec_val is not None and str(spec_val).strip():
            results.append((i + 1, str(spec_val).strip()))
    return results


def parse_spec(raw: str):
    """
    Parse 'T*W*L' → (Decimal(T), Decimal(W), Decimal(L)).
    Returns None and an error message on failure.
    """
    parts = raw.split("*")
    if len(parts) != 3:
        return None, f"Expected 3 segments separated by '*', got {len(parts)}: '{raw}'"
    try:
        thickness = Decimal(parts[0].strip())
        width = Decimal(parts[1].strip())
        length = Decimal(parts[2].strip())
    except InvalidOperation as exc:
        return None, f"Non-numeric segment in '{raw}': {exc}"
    return (thickness, width, length), None


# ------------------------------------------------------------------ #
# Database helpers                                                     #
# ------------------------------------------------------------------ #

def db_get_material(conn, display_name: str):
    """Return item_id or None."""
    with conn.cursor() as cur:
        cur.execute("SELECT id FROM inventory_item WHERE display_name = %s", (display_name,))
        row = cur.fetchone()
    return row[0] if row else None


def db_property_exists(conn, item_id: int, prop_name: str) -> bool:
    with conn.cursor() as cur:
        cur.execute(
            "SELECT 1 FROM item_property WHERE item_id = %s AND property_name = %s",
            (item_id, prop_name),
        )
        return cur.fetchone() is not None


def db_insert_property(conn, item_id: int, prop_name: str, value: Decimal) -> None:
    with conn.cursor() as cur:
        cur.execute(
            """
            INSERT INTO item_property (item_id, property_name, property_value)
            VALUES (%s, %s, %s)
            """,
            (item_id, prop_name, float(value)),
        )
    conn.commit()


# ------------------------------------------------------------------ #
# Per-tab processing                                                   #
# ------------------------------------------------------------------ #

def process_tab(ws, tab_name: str, conn, logger: SeedLogger, retry_keys: set) -> dict:
    """
    Process one tab. Returns per-material stats dict.
    """
    stats = {"tab": tab_name, "specs_found": 0, "specs_parsed": 0, "specs_failed": 0}

    header_idx = find_header_row_idx(ws)
    if header_idx < 0:
        key = f"{tab_name}.all"
        logger.log_failure(key, "Header row not found (no Items + Specifications columns)", f"sheet={tab_name}")
        print(f"  FAIL  [{tab_name}] Header row not found")
        return stats

    display_name = get_display_name_from_tab(ws, header_idx)
    if display_name is None:
        key = f"{tab_name}.all"
        logger.log_failure(key, "Could not derive display_name (Items column empty)", f"sheet={tab_name}")
        print(f"  FAIL  [{tab_name}] Items column empty — cannot identify material")
        return stats

    item_id = db_get_material(conn, display_name)
    if item_id is None:
        key = f"{tab_name}.all"
        logger.log_failure(
            key,
            f"Material '{display_name}' not found in DB — blocked on DE-1.1",
            f"sheet={tab_name}",
        )
        print(f"  BLOCK [{tab_name}] '{display_name}' not in DB — run DE-1.1 first")
        return stats

    # Collect all spec strings in this tab
    spec_rows = collect_spec_strings(ws, header_idx)
    raw_specs = [s for _, s in spec_rows]
    distinct_specs = list(dict.fromkeys(raw_specs))  # preserve insertion order, deduplicate

    stats["specs_found"] = len(distinct_specs)

    if not distinct_specs:
        key = f"{tab_name}.all"
        logger.log_failure(key, "No Specifications values found in tab", f"sheet={tab_name}")
        print(f"  FAIL  [{tab_name}] No Specifications values found")
        return stats

    # Warn about variants (more than one distinct spec)
    if len(distinct_specs) > 1:
        print(f"  WARN  [{tab_name}] {len(distinct_specs)} distinct spec strings found:")
        for spec in distinct_specs:
            print(f"           '{spec}'")

    # Parse the canonical spec (first non-null occurrence)
    canonical_raw, canonical_row = spec_rows[0]
    parsed, parse_err = parse_spec(canonical_raw)

    if parse_err:
        key = f"{tab_name}.parse"
        logger.log_failure(
            key,
            f"Cannot parse canonical spec: {parse_err}",
            f"sheet={tab_name}, row={canonical_row}, raw='{canonical_raw}'",
        )
        print(f"  FAIL  [{tab_name}] {parse_err}")
        stats["specs_failed"] += 1
        return stats

    thickness, width, length = parsed
    stats["specs_parsed"] += 1

    # Also check variant specs for parse errors (informational, don't stop)
    for spec_str in distinct_specs[1:]:
        _, err = parse_spec(spec_str)
        if err:
            key = f"{tab_name}.variant_{spec_str}"
            logger.log_failure(
                key,
                f"Variant spec unparseable: {err}",
                f"sheet={tab_name}, raw='{spec_str}'",
            )
            print(f"  WARN  [{tab_name}] Variant spec failed parse: '{spec_str}' — {err}")
            stats["specs_failed"] += 1

    # Insert each property individually (granular retry tracking)
    prop_values = zip(PROPERTIES, (thickness, width, length))
    for prop_name, prop_value in prop_values:
        retry_key = f"{tab_name}.{prop_name}"

        if retry_keys and retry_key not in retry_keys:
            # In retry mode but this property is not in the failures list — already done.
            continue

        if db_property_exists(conn, item_id, prop_name):
            logger.log_success(retry_key)
            print(f"  SKIP  [{tab_name}] {prop_name}={prop_value} already in DB")
            continue

        try:
            db_insert_property(conn, item_id, prop_name, prop_value)
            logger.log_success(retry_key)
            print(f"  INSERT[{tab_name}] {prop_name}={prop_value} for '{display_name}'")
        except Exception as exc:
            conn.rollback()
            logger.log_failure(
                retry_key,
                f"DB insert failed for {prop_name}: {exc}",
                f"item_id={item_id}, prop_name={prop_name}, value={prop_value}",
            )
            print(f"  FAIL  [{tab_name}] {prop_name}: {exc}")

    return stats


# ------------------------------------------------------------------ #
# Entry point                                                          #
# ------------------------------------------------------------------ #

def main():
    parser = argparse.ArgumentParser(description="DE-1.2: Parse specs and insert dimension properties")
    parser.add_argument("--retry", action="store_true")
    args = parser.parse_args()

    logger = SeedLogger(SCRIPT_NAME)
    failed_keys = logger.get_failed_keys()
    is_retry = args.retry or bool(failed_keys)

    if is_retry:
        print(f"RETRY mode — {len(failed_keys)} property key(s) pending\n")
    else:
        print("FULL mode — processing all valid tabs\n")

    try:
        wb = openpyxl.load_workbook(WORKBOOK_PATH, data_only=True)
    except Exception as exc:
        print(f"ERROR: Cannot open workbook: {exc}", file=sys.stderr)
        sys.exit(1)

    conn = get_connection()

    all_stats = []
    total_failed = 0

    for tab_name in wb.sheetnames:
        if not is_valid_tab(tab_name):
            continue

        # In retry mode, skip tabs that have no pending property keys
        if is_retry:
            tab_keys = {k for k in failed_keys if k.startswith(f"{tab_name}.")}
            if not tab_keys:
                print(f"  SKIP  [{tab_name}] — no pending property keys")
                continue
            retry_keys = tab_keys
        else:
            retry_keys = set()

        ws = wb[tab_name]
        stats = process_tab(ws, tab_name, conn, logger, retry_keys)
        all_stats.append(stats)

    conn.close()

    print("\n" + "=" * 52)
    print("DE-1.2 SUMMARY (per tab)")
    for s in all_stats:
        print(
            f"  {s['tab']:<12} specs_found={s['specs_found']}  "
            f"parsed={s['specs_parsed']}  failed={s['specs_failed']}"
        )

    remaining = logger.get_failed_keys()
    total_failed = len(remaining)
    print(f"\n  Pending failures : {total_failed}")
    if total_failed:
        from seed.seed_logger import FAILURES_LOG
        print(f"  Failure log      : {FAILURES_LOG}")
    print("=" * 52)

    sys.exit(1 if total_failed else 0)


if __name__ == "__main__":
    main()
